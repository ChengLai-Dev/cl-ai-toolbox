# -*- coding: utf-8 -*-

import json
import os
from typing import Any, Dict, List, Optional
from pathlib import Path

from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.types import Tool, TextContent


def _read_xlsx(file_path: str) -> Dict:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=False):
                row_data = []
                for cell in row:
                    row_data.append({
                        "value": cell.value,
                        "column": cell.column,
                        "row": cell.row,
                        "coordinate": cell.coordinate
                    })
                rows.append(row_data)
            sheets[sheet_name] = {
                "rows": rows,
                "max_row": ws.max_row,
                "max_column": ws.max_column
            }
        wb.close()
        return {
            "success": True,
            "sheets": sheets,
            "sheet_names": wb.sheetnames,
            "path": file_path
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def _read_xls(file_path: str) -> Dict:
    try:
        import xlrd
        wb = xlrd.open_workbook(file_path)
        sheets = {}
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            rows = []
            for r in range(ws.nrows):
                row_data = []
                for c in range(ws.ncols):
                    row_data.append({
                        "value": ws.cell_value(r, c),
                        "column": c + 1,
                        "row": r + 1,
                        "coordinate": f"{chr(65 + c)}{r + 1}"
                    })
                rows.append(row_data)
            sheets[sheet_name] = {
                "rows": rows,
                "max_row": ws.nrows,
                "max_column": ws.ncols
            }
        return {
            "success": True,
            "sheets": sheets,
            "sheet_names": wb.sheet_names(),
            "path": file_path
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def _write_xlsx(file_path: str, data: Dict[str, List[List[Any]]]) -> Dict:
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sheet_name, rows in data.items():
            ws = wb.create_sheet(title=sheet_name)
            for r, row_data in enumerate(rows, 1):
                for c, value in enumerate(row_data, 1):
                    ws.cell(row=r, column=c, value=value)
        os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)
        wb.save(file_path)
        wb.close()
        return {"success": True, "message": f"Excel 文件已保存到: {file_path}", "path": file_path}
    except Exception as e:
        return {"success": False, "error": str(e)}


def _write_xls(file_path: str, data: Dict[str, List[List[Any]]]) -> Dict:
    try:
        import xlwt
        wb = xlwt.Workbook()
        for sheet_name, rows in data.items():
            ws = wb.add_sheet(sheet_name)
            for r, row_data in enumerate(rows):
                for c, value in enumerate(row_data):
                    ws.write(r, c, value)
        os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)
        wb.save(file_path)
        return {"success": True, "message": f"Excel 文件已保存到: {file_path}", "path": file_path}
    except Exception as e:
        return {"success": False, "error": str(e)}


def _update_cell_xlsx(file_path: str, sheet_name: str, row: int, column: int, value: Any) -> Dict:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"工作表 '{sheet_name}' 不存在"}
        ws = wb[sheet_name]
        ws.cell(row=row, column=column, value=value)
        wb.save(file_path)
        wb.close()
        return {"success": True, "message": f"单元格 {chr(64 + column)}{row} 已更新", "path": file_path}
    except Exception as e:
        return {"success": False, "error": str(e)}


server = Server("excel-mcp")


@server.list_tools()
async def list_tools() -> List[Tool]:
    return [
        Tool(
            name="read_excel",
            description="读取 Excel 文件内容（支持 .xlsx 和 .xls 格式），返回所有工作表数据",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件的绝对路径"
                    }
                },
                "required": ["file_path"]
            }
        ),
        Tool(
            name="create_excel",
            description="创建新的 Excel 文件，写入指定数据",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "要创建的 Excel 文件路径（.xlsx 或 .xls）"
                    },
                    "data": {
                        "type": "object",
                        "description": "工作表数据，键为工作表名，值为二维数组（行->列）",
                        "additionalProperties": {
                            "type": "array",
                            "items": {
                                "type": "array",
                                "items": {}
                            }
                        }
                    }
                },
                "required": ["file_path", "data"]
            }
        ),
        Tool(
            name="update_cell",
            description="更新 Excel 文件中指定单元格的值（仅支持 .xlsx）",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": ".xlsx 文件路径"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称"
                    },
                    "row": {
                        "type": "integer",
                        "description": "行号（从1开始）"
                    },
                    "column": {
                        "type": "integer",
                        "description": "列号（从1开始）",
                        "default": 1
                    },
                    "value": {
                        "description": "要设置的值"
                    }
                },
                "required": ["file_path", "sheet_name", "row", "column", "value"]
            }
        )
    ]


@server.call_tool()
async def call_tool(name: str, arguments: Dict[str, Any]) -> List[TextContent]:
    result = {}

    if name == "read_excel":
        file_path = arguments.get("file_path", "")
        if not file_path or not os.path.exists(file_path):
            result = {"success": False, "error": f"文件不存在: {file_path}"}
        else:
            ext = os.path.splitext(file_path)[1].lower()
            if ext == ".xlsx":
                result = _read_xlsx(file_path)
            elif ext == ".xls":
                result = _read_xls(file_path)
            else:
                result = {"success": False, "error": f"不支持的文件格式: {ext}，仅支持 .xls 和 .xlsx"}

    elif name == "create_excel":
        file_path = arguments.get("file_path", "")
        data = arguments.get("data", {})
        if not file_path or not data:
            result = {"success": False, "error": "文件路径和数据不能为空"}
        else:
            ext = os.path.splitext(file_path)[1].lower()
            if ext == ".xlsx":
                result = _write_xlsx(file_path, data)
            elif ext == ".xls":
                result = _write_xls(file_path, data)
            else:
                result = {"success": False, "error": f"不支持的文件格式: {ext}，仅支持 .xls 和 .xlsx"}

    elif name == "update_cell":
        file_path = arguments.get("file_path", "")
        sheet_name = arguments.get("sheet_name", "")
        row = arguments.get("row", 0)
        column = arguments.get("column", 1)
        value = arguments.get("value")
        if not file_path or not sheet_name or row < 1:
            result = {"success": False, "error": "文件路径、工作表名和行号不能为空"}
        elif not os.path.exists(file_path):
            result = {"success": False, "error": f"文件不存在: {file_path}"}
        else:
            ext = os.path.splitext(file_path)[1].lower()
            if ext == ".xlsx":
                result = _update_cell_xlsx(file_path, sheet_name, row, column, value)
            else:
                result = {"success": False, "error": "仅支持 .xlsx 格式的单元格更新"}

    else:
        result = {"success": False, "error": f"未知的工具: {name}"}

    return [TextContent(type="text", text=json.dumps(result, ensure_ascii=False, indent=2))]


async def main():
    async with stdio_server() as (read_stream, write_stream):
        await server.run(read_stream, write_stream, server.create_initialization_options())


if __name__ == "__main__":
    import asyncio
    asyncio.run(main())
